import openpyxl
import os
import sys
from datetime import datetime, date, time
import uuid
import random
import telebot
from telebot import types

print("Bot is running")

bot = telebot.TeleBot('')

admin="0"
ausername="0"

z=0

comp1="0"
series1="0"
sock3="0"
price2="0"
price3="0"
wf3="0"
user="0"
rusername="0"

answ4=["HYPERX Fury Black; DDR4; 2666MHz; 8GB(Kit 2x4GB); Цена: +-70$",
       "HYPERX Fury Black; DDR4; 2666MHz; 16GB(Kit 2x8GB); Цена: +-90$",
       "HYPERX Predator Black; DDR4; 3200MHz; 16GB(Kit 2x8GB); Цена: +-110$",
       "HYPERX Predator Black; DDR4; 3200MHz; 32GB(Kit 2x16GB); Цена: +-200$",
       "HYPERX Predator Black; DDR4; 3200MHz; 64GB(Kit 2x32GB); Цена: +-350$"]
answ5=["DEEPCOOL Gammaxx 300; Цена: +-20$",
       "DEEPCOOL Gammaxx 300; Цена: +-20$",
       "DEEPCOOL Gammaxx 300; Цена: +-20$",
       "DEEPCOOL Gammaxx 300; Цена: +-20$",
       "THERMALRIGHT Silver Arrow T8; Цена: +-95$"]
answ6=["SSD ADATA XPG SX8100; 256GB; Цена: +-55$",
       "SSD ADATA XPG SX8200 Pro; 512GB; Цена: +-75$",
       "SSD ADATA XPG Spectrix S40G; 512GB; Цена: +-100$",
       "SSD ADATA XPG SX8200 Pro; 1TB; Цена: +-170$",
       "SD ADATA XPG Spectrix S40G; 4TB; Цена: +-600$"]
answ7="0"
answ8=["500W GAMEMAX GP-500; Цена: +-40$",
       "600W GAMEMAX GM-600; Цена: +-50$",
       "700W GAMEMAX GM-700; Цена: +-70$",
       "750W DEEPCOOL DQ750ST; Цена: +-90$",
       "1600W CORSAIR AX1600i; Цена: +-660$"]
answend="Я надеюсь я тебе помог :), надо будет помощь ещё, напиши /start"
reviewansw="Если ты хочешь оставить отзыв/пожелание моему автору пропиши команду /review"

def filtercpu():
    #CPU------------------------------------------------------------------------------------------

    global sock3

    filtrfilename=uuid.uuid4().hex+".xlsx"

    wb1=openpyxl.reader.excel.load_workbook(filename="CPU.xlsx", data_only=True)
    wb1.active=0
    ws1=wb1.active

    rfilt1=[]
    k=2
    for a in range(2, ws1.max_row+1):
        if ws1['A'+str(a)].value == comp1:
            rfilt1.append(k)
        k+=1

    if len(rfilt1)==0:
        os.remove(filtrfilename)
        sys.exit("No matches detected")

    wbf1=openpyxl.Workbook()
    wsf1=wbf1.active

    for b in range(1, len(rfilt1)+1):
        wsf1['A'+str(b)] = ws1['A'+str(rfilt1[b-1])].value
        wsf1['B'+str(b)] = ws1['B'+str(rfilt1[b-1])].value
        wsf1['C'+str(b)] = ws1['C'+str(rfilt1[b-1])].value
        wsf1['D'+str(b)] = ws1['D'+str(rfilt1[b-1])].value
        wsf1['E'+str(b)] = ws1['E'+str(rfilt1[b-1])].value
        wsf1['F'+str(b)] = ws1['F'+str(rfilt1[b-1])].value
        wsf1['G'+str(b)] = ws1['G'+str(rfilt1[b-1])].value
        wsf1['H'+str(b)] = ws1['H'+str(rfilt1[b-1])].value
        wsf1['I'+str(b)] = ws1['I'+str(rfilt1[b-1])].value
        wsf1['J'+str(b)] = ws1['J'+str(rfilt1[b-1])].value

    wbf1.save(filtrfilename)

    wbf2=openpyxl.Workbook()
    wbf2.active=0
    wsf2=wbf2.active

    rfilt2=[]
    l=1
    for c in range(1, wsf1.max_row+1):
        if wsf1['C'+str(c)].value == series1:
            rfilt2.append(l)
        l+=1

    if len(rfilt2)==0:
        os.remove(filtrfilename)
        sys.exit("No matches detected")


    for d in range(1, len(rfilt2)+1):
        wsf2['A'+str(d)] = wsf1['A'+str(rfilt2[d-1])].value
        wsf2['B'+str(d)] = wsf1['B'+str(rfilt2[d-1])].value
        wsf2['C'+str(d)] = wsf1['C'+str(rfilt2[d-1])].value
        wsf2['D'+str(d)] = wsf1['D'+str(rfilt2[d-1])].value
        wsf2['E'+str(d)] = wsf1['E'+str(rfilt2[d-1])].value
        wsf2['F'+str(d)] = wsf1['F'+str(rfilt2[d-1])].value
        wsf2['G'+str(d)] = wsf1['G'+str(rfilt2[d-1])].value
        wsf2['H'+str(d)] = wsf1['H'+str(rfilt2[d-1])].value
        wsf2['I'+str(d)] = wsf1['I'+str(rfilt2[d-1])].value
        wsf2['J'+str(d)] = wsf1['J'+str(rfilt2[d-1])].value

    wbf2.save(filtrfilename)

    wb2=openpyxl.reader.excel.load_workbook(filename=filtrfilename, data_only=True)
    wb2.active=0
    ws2=wb2.active

    for e in range(1, ws2.max_row+1):
        print(ws2['A'+str(e)].value,
              ws2['B'+str(e)].value,
              ws2['C'+str(e)].value,
              ws2['D'+str(e)].value, ";",
              ws2['E'+str(e)].value, ";",
              ws2['F'+str(e)].value, ";",
              ws2['G'+str(e)].value, ";",
              ws2['H'+str(e)].value, ";",
              ws2['I'+str(e)].value, ";",
              ws2['J'+str(e)].value)

    r=random.randint(1, ws2.max_row)
    answ1="Процессор: "+str(ws2['A'+str(r)].value)+" "+str(ws2['B'+str(r)].value)+" "+str(ws2['C'+str(r)].value)+" "+str(ws2['D'+str(r)].value)+"; "+str(ws2['E'+str(r)].value)+"; "+str(ws2['F'+str(r)].value)+"; "+str(ws2['G'+str(r)].value)+"; "+str(ws2['H'+str(r)].value)+"; "+str(ws2['I'+str(r)].value)+"; "+"Цена: "+str(ws2['J'+str(r)].value)

    os.remove(filtrfilename)

    sock3=str(ws2['G'+str(r)].value)

    return answ1

def filtercpu2():
    #CPU------------------------------------------------------------------------------------------

    global sock3

    filtrfilename=uuid.uuid4().hex+".xlsx"

    wb1=openpyxl.reader.excel.load_workbook(filename="CPU.xlsx", data_only=True)
    wb1.active=0
    ws1=wb1.active

    rfilt1=[]
    k=2
    for a in range(2, ws1.max_row+1):
        if ws1['A'+str(a)].value == comp1:
            rfilt1.append(k)
        k+=1

    if len(rfilt1)==0:
        os.remove(filtrfilename)
        sys.exit("No matches detected")

    wbf1=openpyxl.Workbook()
    wsf1=wbf1.active

    for b in range(1, len(rfilt1)+1):
        wsf1['A'+str(b)] = ws1['A'+str(rfilt1[b-1])].value
        wsf1['B'+str(b)] = ws1['B'+str(rfilt1[b-1])].value
        wsf1['C'+str(b)] = ws1['C'+str(rfilt1[b-1])].value
        wsf1['D'+str(b)] = ws1['D'+str(rfilt1[b-1])].value
        wsf1['E'+str(b)] = ws1['E'+str(rfilt1[b-1])].value
        wsf1['F'+str(b)] = ws1['F'+str(rfilt1[b-1])].value
        wsf1['G'+str(b)] = ws1['G'+str(rfilt1[b-1])].value
        wsf1['H'+str(b)] = ws1['H'+str(rfilt1[b-1])].value
        wsf1['I'+str(b)] = ws1['I'+str(rfilt1[b-1])].value
        wsf1['J'+str(b)] = ws1['J'+str(rfilt1[b-1])].value

    wbf1.save(filtrfilename)

    wbf2=openpyxl.Workbook()
    wbf2.active=0
    wsf2=wbf2.active

    rfilt2=[]
    l=1
    for c in range(1, wsf1.max_row+1):
        if wsf1['C'+str(c)].value == series1:
            rfilt2.append(l)
        l+=1

    if len(rfilt2)==0:
        os.remove(filtrfilename)
        sys.exit("No matches detected")


    for d in range(1, len(rfilt2)+1):
        wsf2['A'+str(d)] = wsf1['A'+str(rfilt2[d-1])].value
        wsf2['B'+str(d)] = wsf1['B'+str(rfilt2[d-1])].value
        wsf2['C'+str(d)] = wsf1['C'+str(rfilt2[d-1])].value
        wsf2['D'+str(d)] = wsf1['D'+str(rfilt2[d-1])].value
        wsf2['E'+str(d)] = wsf1['E'+str(rfilt2[d-1])].value
        wsf2['F'+str(d)] = wsf1['F'+str(rfilt2[d-1])].value
        wsf2['G'+str(d)] = wsf1['G'+str(rfilt2[d-1])].value
        wsf2['H'+str(d)] = wsf1['H'+str(rfilt2[d-1])].value
        wsf2['I'+str(d)] = wsf1['I'+str(rfilt2[d-1])].value
        wsf2['J'+str(d)] = wsf1['J'+str(rfilt2[d-1])].value

    wbf2.save(filtrfilename)

    wbf21=openpyxl.Workbook()
    wbf21.active=0
    wsf21=wbf21.active

    rfilt21=[]
    l=1
    for c in range(1, wsf2.max_row+1):
        if wsf2['H'+str(c)].value != graph1:
            rfilt21.append(l)
        l+=1

    if len(rfilt21)==0:
        os.remove(filtrfilename)
        sys.exit("No matches detected")


    for f in range(1, len(rfilt21)+1):
        wsf21['A'+str(f)] = wsf2['A'+str(rfilt21[f-1])].value
        wsf21['B'+str(f)] = wsf2['B'+str(rfilt21[f-1])].value
        wsf21['C'+str(f)] = wsf2['C'+str(rfilt21[f-1])].value
        wsf21['D'+str(f)] = wsf2['D'+str(rfilt21[f-1])].value
        wsf21['E'+str(f)] = wsf2['E'+str(rfilt21[f-1])].value
        wsf21['F'+str(f)] = wsf2['F'+str(rfilt21[f-1])].value
        wsf21['G'+str(f)] = wsf2['G'+str(rfilt21[f-1])].value
        wsf21['H'+str(f)] = wsf2['H'+str(rfilt21[f-1])].value
        wsf21['I'+str(f)] = wsf2['I'+str(rfilt21[f-1])].value
        wsf21['J'+str(f)] = wsf2['J'+str(rfilt21[f-1])].value

    wbf21.save(filtrfilename)

    wb2=openpyxl.reader.excel.load_workbook(filename=filtrfilename, data_only=True)
    wb2.active=0
    ws2=wb2.active

    for e in range(1, wsf21.max_row+1):
        print(ws2['A'+str(e)].value,
              ws2['B'+str(e)].value,
              ws2['C'+str(e)].value,
              ws2['D'+str(e)].value, ";",
              ws2['E'+str(e)].value, ";",
              ws2['F'+str(e)].value, ";",
              ws2['G'+str(e)].value, ";",
              ws2['H'+str(e)].value, ";",
              ws2['I'+str(e)].value, ";",
              ws2['J'+str(e)].value)

    r=random.randint(1, wsf21.max_row)
    answ1="Процессор: "+str(ws2['A'+str(r)].value)+" "+str(ws2['B'+str(r)].value)+" "+str(ws2['C'+str(r)].value)+" "+str(ws2['D'+str(r)].value)+"; "+str(ws2['E'+str(r)].value)+"; "+str(ws2['F'+str(r)].value)+"; "+str(ws2['G'+str(r)].value)+"; "+str(ws2['H'+str(r)].value)+"; "+str(ws2['I'+str(r)].value)+"; "+"Цена: "+str(ws2['J'+str(r)].value)

    sock3=str(ws2['G'+str(r)].value)

    os.remove(filtrfilename)

    return answ1

def filtermb1():
    #MB-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    filtrfilename=uuid.uuid4().hex+".xlsx"

    wb3=openpyxl.reader.excel.load_workbook(filename="MB.xlsx", data_only=True)
    wb3.active=0
    ws3=wb3.active

    rfilt3=[]
    k=2
    for a in range(2, ws3.max_row):
        if ws3['C'+str(a)].value == sock3:
            rfilt3.append(k)
        k+=1

    if len(rfilt3)==0:
        os.remove(filtrfilename)
        sys.exit("No matches detected")

    wbf3=openpyxl.Workbook()
    wsf3=wbf3.active

    for b in range(1, len(rfilt3)+1):
        wsf3['A'+str(b)] = ws3['A'+str(rfilt3[b-1])].value
        wsf3['B'+str(b)] = ws3['B'+str(rfilt3[b-1])].value
        wsf3['C'+str(b)] = ws3['C'+str(rfilt3[b-1])].value
        wsf3['D'+str(b)] = ws3['D'+str(rfilt3[b-1])].value
        wsf3['E'+str(b)] = ws3['E'+str(rfilt3[b-1])].value

    wbf3.save(filtrfilename)

    wbf4=openpyxl.Workbook()
    wbf4.active=0
    wsf4=wbf4.active

    rfilt4=[]
    l=1
    for c in range(1, wsf3.max_row+1):
        if wsf3['D'+str(c)].value == wf3:
            rfilt4.append(l)
        l+=1

    if len(rfilt4)==0:
        os.remove(filtrfilename)
        sys.exit("No matches detected")

    for d in range(1, len(rfilt4)+1):
        wsf4['A'+str(d)] = wsf3['A'+str(rfilt4[d-1])].value
        wsf4['B'+str(d)] = wsf3['B'+str(rfilt4[d-1])].value
        wsf4['C'+str(d)] = wsf3['C'+str(rfilt4[d-1])].value
        wsf4['D'+str(d)] = wsf3['D'+str(rfilt4[d-1])].value
        wsf4['E'+str(d)] = wsf3['E'+str(rfilt4[d-1])].value

    wbf4.save(filtrfilename)

    wbf41=openpyxl.Workbook()
    wbf41.active=0
    wsf41=wbf41.active

    rfilt41=[]
    l=1
    for c in range(1, wsf4.max_row+1):
        if wsf4['E'+str(c)].value == price3:
            rfilt41.append(l)
        l+=1

    if len(rfilt41)==0:
        os.remove(filtrfilename)
        sys.exit("No matches detected")

    for f in range(1, len(rfilt41)+1):
        wsf41['A'+str(f)] = wsf4['A'+str(rfilt41[f-1])].value
        wsf41['B'+str(f)] = wsf4['B'+str(rfilt41[f-1])].value
        wsf41['C'+str(f)] = wsf4['C'+str(rfilt41[f-1])].value
        wsf41['D'+str(f)] = wsf4['D'+str(rfilt41[f-1])].value
        wsf41['E'+str(f)] = wsf4['E'+str(rfilt41[f-1])].value


    wbf41.save(filtrfilename)

    wb4=openpyxl.reader.excel.load_workbook(filename=filtrfilename, data_only=True)
    wb4.active=0
    ws4=wb4.active

    for e in range(1, wsf41.max_row+1):
        print(ws4['A'+str(e)].value,
              ws4['B'+str(e)].value,
              ws4['C'+str(e)].value,
              ws4['D'+str(e)].value, ";",
              ws4['E'+str(e)].value, ";")

    r=random.randint(1, wsf41.max_row)
    answ2="Материнская плата: "+str(ws4['A'+str(r)].value)+" "+str(ws4['B'+str(r)].value)+"; "+str(ws4['C'+str(r)].value)+"; "+str(ws4['D'+str(r)].value)+"; "+"Цена: "+str(ws4['E'+str(r)].value)

    os.remove(filtrfilename)

    return answ2

def filtergpu():
    #GPU------------------------------------------------------------------------------------------

    filtrfilename=uuid.uuid4().hex+".xlsx"

    wb5=openpyxl.reader.excel.load_workbook(filename="GPU.xlsx", data_only=True)
    wb5.active=0
    ws5=wb5.active

    rfilt5=[]
    k=2
    for a in range(2, ws5.max_row+1):
        if ws5['F'+str(a)].value == price2:
            rfilt5.append(k)
        k+=1

    if len(rfilt5)==0:
        os.remove(filtrfilename)
        sys.exit("No matches detected")

    wbf5=openpyxl.Workbook()
    wbf5.active=0
    wsf5=wbf5.active

    for b in range(1, len(rfilt5)+1):
        wsf5['A'+str(b)] = ws5['A'+str(rfilt5[b-1])].value
        wsf5['B'+str(b)] = ws5['B'+str(rfilt5[b-1])].value
        wsf5['C'+str(b)] = ws5['C'+str(rfilt5[b-1])].value
        wsf5['D'+str(b)] = ws5['D'+str(rfilt5[b-1])].value
        wsf5['E'+str(b)] = ws5['E'+str(rfilt5[b-1])].value
        wsf5['F'+str(b)] = ws5['F'+str(rfilt5[b-1])].value
        wsf5['G'+str(b)] = ws5['G'+str(rfilt5[b-1])].value
        wsf5['H'+str(b)] = ws5['H'+str(rfilt5[b-1])].value
        wsf5['I'+str(b)] = ws5['I'+str(rfilt5[b-1])].value
        wsf5['J'+str(b)] = ws5['J'+str(rfilt5[b-1])].value

    wbf5.save(filtrfilename)

    wb6=openpyxl.reader.excel.load_workbook(filename=filtrfilename, data_only=True)
    wb6.active=0
    ws6=wb6.active

    for e in range(1, ws6.max_row+1):
        print(ws6['A'+str(e)].value,
              ws6['B'+str(e)].value,
              ws6['C'+str(e)].value, ";",
              ws6['D'+str(e)].value, ";",
              ws6['E'+str(e)].value, ";",
              ws6['F'+str(e)].value)

    n=random.randint(1, ws6.max_row)
    answ3="Видеокарта: "+str(ws6['A'+str(n)].value)+" "+str(ws6['B'+str(n)].value)+" "+str(ws6['C'+str(n)].value)+"; "+str(ws6['D'+str(n)].value)+"; "+str(ws6['E'+str(n)].value)+"; "+str(ws6['F'+str(n)].value)

    os.remove(filtrfilename)

    return answ3

def filterbox():
    #BOX-------------------------------------------------------------------------

    wb7=openpyxl.reader.excel.load_workbook(filename="BOX.xlsx", data_only=True)
    wb7.active=0
    ws7=wb7.active

    n=random.randint(1, ws7.max_row)
    answ7="Корпус: "+str(ws7['A'+str(n)].value)+"; Цена: +-50$"
    print(answ7)

    return answ7

#Bot-----------------------------------------------------------------------------

#admin---------------------------------------------------------------------------
@bot.message_handler(commands=['admin'])
def admin_message(message):
    global admin
    sent=bot.send_message(message.from_user.id, "Введи пароль")
    bot.register_next_step_handler(sent, admin_password)

def admin_password(message):
    global admin
    global ausername
    if message.text == "01gleb09":
        admin=message.from_user.id
        print(admin)
        if bool(message.from_user.username)==True:
            ausername="@"+message.from_user.username
        else:
            ausername='"'+message.from_user.first_name+'"'
        bot.send_message(admin, "Ты теперь админ")
    else:
        bot.send_message(message.from_user.id, "Пароль неверный")
        bot.send_message(message.from_user.id, "Ты не админ")

@bot.message_handler(commands=['adminn'])
def admin_message(message):
    global admin
    global ausername
    if admin=="0":
        bot.send_message(message.from_user.id, "Админ не зарегистрирован")
    else:
        bot.send_message(message.from_user.id, ausername)
#general--------------------------------------------------------------------------

#answforming----------------------------------------------------------------------
def answ_message(call):
    global comp1
    global series1
    global graph1
    global sock3
    global wf3
    global price2
    global price3
    global answ4
    global answ5
    global answ6
    global answ7
    global answ8
    global answend
    global reviewansw
    global z
    bot.send_message(call.message.chat.id, "Кхм...")
    bot.send_message(call.message.chat.id, "Рекомедую:")
    bot.send_message(call.message.chat.id, filtercpu())
    bot.send_message(call.message.chat.id, filtermb1())
    bot.send_message(call.message.chat.id, filtergpu())
    bot.send_message(call.message.chat.id, "Оперативная память: "+answ4[z])
    bot.send_message(call.message.chat.id, "Охлаждение процессора(если в комплекте с процессором нету): "+answ5[z])
    bot.send_message(call.message.chat.id, "Постоянная память: "+answ6[z])
    bot.send_message(call.message.chat.id, filterbox())
    bot.send_message(call.message.chat.id, "Блок питания: "+answ8[z])
    bot.send_message(call.message.chat.id, answend)
    bot.send_message(call.message.chat.id, reviewansw)

def answokfunc():
    n=random.randint(1, 5)
    if n==1:
        answok="Хорошо, следуйщий вопрос"
    elif n==2:
        answok="Записал"
    elif n==3:
        answok="Запомнил"
    elif n==4:
        answok="Принял"
    elif n==5:
        answok="Хороший выбор"
    return answok
#commands------------------------------------------------------------------------
@bot.message_handler(commands=['help'])
def help_messaage(message):
    bot.send_message(message.from_user.id, "/start - начать сборку")
    bot.send_message(message.from_user.id, "/review - передать отзыв")
    bot.send_message(message.from_user.id, "/help - помощь")
    bot.send_message(message.from_user.id, "/about - об о мне")

@bot.message_handler(commands=['about'])
def about_messaage(message):
    bot.send_message(message.from_user.id, "Я бот-конфигуратор, я был создан для того что б помогать людям выбирать комплектующие для ПК")
#review--------------------------------------------------------------------------
@bot.message_handler(commands=['review'])
def review_message(message):
    global admin
    global rusername
    if bool(message.from_user.username)==True:
        rusername="@"+message.from_user.username
    else:
        rusername='"'+message.from_user.first_name+'"'
    sent=bot.send_message(message.from_user.id, "Напиши одним сообщением что ты хочешь что б я передал")
    bot.register_next_step_handler(sent, review_text)

def review_text(message):
    global admin
    global rusername
    bot.send_message(admin, "Новый отзыв от "+rusername+":")
    bot.send_message(admin, message.text)
    bot.send_message(message.from_user.id, "Передал:)")
#generalwork----------------------------------------------------------------------
@bot.message_handler(commands=['start'])
def welcome_message(message):
    user = bot.get_me()
    print("____Time____: "+str(datetime.today()))
    print("____Bot____: "+str(user))
    print("____User____: "+str(message))
    bot.send_message(message.from_user.id, "Привет")
    keyboard = types.InlineKeyboardMarkup()
    key_yes = types.InlineKeyboardButton(text='Да', callback_data='yes')
    keyboard.add(key_yes)
    key_no= types.InlineKeyboardButton(text='Нет', callback_data='no')
    keyboard.add(key_no)
    question = 'Ты хочешь собрать ПК?'
    bot.send_message(message.from_user.id, text=question, reply_markup=keyboard)

@bot.callback_query_handler(func=lambda call: True)
def callback_worker(call):
    global comp1
    global series1
    global graph1
    global sock3
    global wf3
    global price2
    global price3
    global answ4
    global answ5
    global answ6
    global answ7
    global answ8
    global answend
    global reviewansw
    global z
    if call.data == "yes":
        bot.send_message(call.message.chat.id, "Тогда ты попал по адресу :)")
        bot.send_message(call.from_user.id, "Я задам тебе пару вопросов, что б понять что тебе подбирать")
        keyboard = types.InlineKeyboardMarkup()
        key_game = types.InlineKeyboardButton(text='Игровой', callback_data='game')
        keyboard.add(key_game)
        key_work= types.InlineKeyboardButton(text='Рабочий', callback_data='work')
        keyboard.add(key_work)
        question = 'Ты собираешь ... ПК'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "no":
        bot.send_message(call.message.chat.id, 'Жалко :(')
        bot.send_message(call.from_user.id, "Если надумаешь приходи и напиши /start")
    elif call.data == "game":
        bot.send_message(call.message.chat.id, answokfunc())
        keyboard = types.InlineKeyboardMarkup()
        key_amd = types.InlineKeyboardButton(text='AMD', callback_data='amd')
        keyboard.add(key_amd)
        key_intel= types.InlineKeyboardButton(text='INTEL', callback_data='intel')
        keyboard.add(key_intel)
        key_netral= types.InlineKeyboardButton(text='Неважно/Не знаю', callback_data='netral')
        keyboard.add(key_netral)
        question = 'Выбери:'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "amd":
        comp1="AMD"
        bot.send_message(call.message.chat.id, answokfunc())
        keyboard = types.InlineKeyboardMarkup()
        key_lite = types.InlineKeyboardButton(text='<500$', callback_data='lite')
        keyboard.add(key_lite)
        key_norm= types.InlineKeyboardButton(text='500-1500$', callback_data='norm')
        keyboard.add(key_norm)
        key_much = types.InlineKeyboardButton(text='1500-2500$', callback_data='much')
        keyboard.add(key_much)
        key_toomuch = types.InlineKeyboardButton(text='>2500$', callback_data='toomuch')
        keyboard.add(key_toomuch)
        key_inf = types.InlineKeyboardButton(text='+-∞$', callback_data='inf')
        keyboard.add(key_inf)
        question = 'Какой бюджет?'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "intel":
        comp1="INTEL"
        bot.send_message(call.message.chat.id, answokfunc())
        keyboard = types.InlineKeyboardMarkup()
        key_lite = types.InlineKeyboardButton(text='<500$', callback_data='lite')
        keyboard.add(key_lite)
        key_norm= types.InlineKeyboardButton(text='500-1500$', callback_data='norm')
        keyboard.add(key_norm)
        key_much = types.InlineKeyboardButton(text='1500-2500$', callback_data='much')
        keyboard.add(key_much)
        key_toomuch = types.InlineKeyboardButton(text='>2500$', callback_data='toomuch')
        keyboard.add(key_toomuch)
        key_inf = types.InlineKeyboardButton(text='+-∞$', callback_data='inf')
        keyboard.add(key_inf)
        question = 'Какой бюджет?'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "netral":
        n=random.randint(1, 2)
        if n==1:
            comp1="AMD"
        else:
            comp1="INTEL"
        bot.send_message(call.message.chat.id, answokfunc())
        keyboard = types.InlineKeyboardMarkup()
        key_lite = types.InlineKeyboardButton(text='<500$', callback_data='lite')
        keyboard.add(key_lite)
        key_norm= types.InlineKeyboardButton(text='500-1500$', callback_data='norm')
        keyboard.add(key_norm)
        key_much = types.InlineKeyboardButton(text='1500-2500$', callback_data='much')
        keyboard.add(key_much)
        key_toomuch = types.InlineKeyboardButton(text='>2500$', callback_data='toomuch')
        keyboard.add(key_toomuch)
        key_inf = types.InlineKeyboardButton(text='+-∞$', callback_data='inf')
        keyboard.add(key_inf)
        question = 'Какой бюджет?'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "lite":
        if comp1=="AMD":
            series1="3"
        else:
            series1="i3"
        price2="100-250$"
        price3="<100$"
        wf3="Wi-Fi/Bluetooth – отсутствуют"
        z=0
        answ_message(call)
    elif call.data == "norm":
        bot.send_message(call.message.chat.id, answokfunc())
        if comp1=="AMD":
            series1="5"
        else:
            series1="i5"
        price2="250-500$"
        price3="100-200$"
        z=1
        keyboard = types.InlineKeyboardMarkup()
        key_yes3 = types.InlineKeyboardButton(text='Да', callback_data='yes3')
        keyboard.add(key_yes3)
        key_no3= types.InlineKeyboardButton(text='Нет', callback_data='no3')
        keyboard.add(key_no3)
        question = 'Нужен Wi-Fi/Bluetooth?'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "much":
        bot.send_message(call.message.chat.id, answokfunc())
        if comp1=="AMD":
            series1="7"
        else:
            series1="i7"
        price2="500-1000$"
        price3="200-300$"
        z=2
        keyboard = types.InlineKeyboardMarkup()
        key_yes3 = types.InlineKeyboardButton(text='Да', callback_data='yes3')
        keyboard.add(key_yes3)
        key_no3= types.InlineKeyboardButton(text='Нет', callback_data='no3')
        keyboard.add(key_no3)
        question = 'Нужен Wi-Fi/Bluetooth?'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "toomuch":
        bot.send_message(call.message.chat.id, answokfunc())
        if comp1=="AMD":
            series1="9"
        else:
            series1="i9"
        price2=">1000$"
        price3="200-300$"
        z=3
        keyboard = types.InlineKeyboardMarkup()
        key_yes3 = types.InlineKeyboardButton(text='Да', callback_data='yes3')
        keyboard.add(key_yes3)
        key_no3= types.InlineKeyboardButton(text='Нет', callback_data='no3')
        keyboard.add(key_no3)
        question = 'Нужен Wi-Fi/Bluetooth?'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "inf":
        bot.send_message(call.message.chat.id, answokfunc())
        if comp1=="AMD":
            series1="Threadripper"
        else:
            series1="W"
        price2=">6000$"
        price3=">500$"
        wf3="Wi-Fi/Bluetooth – присутствуют"
        z=4
        answ_message(call)
    elif call.data == "work":
        bot.send_message(call.message.chat.id, answokfunc())
        keyboard = types.InlineKeyboardMarkup()
        key_yes = types.InlineKeyboardButton(text='Да', callback_data='yes2')
        keyboard.add(key_yes)
        key_no= types.InlineKeyboardButton(text='Нет', callback_data='no2')
        keyboard.add(key_no)
        question = 'Работа для ПК связана с графикой?'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "yes2":
        bot.send_message(call.message.chat.id, answokfunc())
        comp1="INTEL"
        graph1="без встроеной графики"
        keyboard = types.InlineKeyboardMarkup()
        key_lite2 = types.InlineKeyboardButton(text='<500$', callback_data='lite2')
        keyboard.add(key_lite2)
        key_norm2= types.InlineKeyboardButton(text='500-1500$', callback_data='norm2')
        keyboard.add(key_norm2)
        key_much2 = types.InlineKeyboardButton(text='1500-2500$', callback_data='much2')
        keyboard.add(key_much2)
        key_toomuch2 = types.InlineKeyboardButton(text='>2500$', callback_data='toomuch2')
        keyboard.add(key_toomuch2)
        question = 'Какой бюджет?'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "no2":
        bot.send_message(call.message.chat.id, answokfunc())
        comp1="INTEL"
        price2="<100$"
        graph1="без встроеной графики"
        keyboard = types.InlineKeyboardMarkup()
        key_lite2 = types.InlineKeyboardButton(text='<500$', callback_data='lite2')
        keyboard.add(key_lite2)
        key_norm2 = types.InlineKeyboardButton(text='500-1500$', callback_data='norm2')
        keyboard.add(key_norm2)
        key_much2 = types.InlineKeyboardButton(text='1500-2500$', callback_data='much2')
        keyboard.add(key_much2)
        key_toomuch2 = types.InlineKeyboardButton(text='>2500$', callback_data='toomuch2')
        keyboard.add(key_toomuch2)
        question = 'Какой бюджет?'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "lite2":
        bot.send_message(call.message.chat.id, answokfunc())
        if price2 !="<100$":
            price2="100-250$"
        series1="i3"
        price3="<100$"
        wf3="Wi-Fi/Bluetooth – отсутствуют"
        z=0
        answ_message(call)
    elif call.data == "norm2":
        bot.send_message(call.message.chat.id, answokfunc())
        if price2 !="<100$":
            price2="250-500$"
        series1="i5"
        price3="100-200$"
        z=1
        keyboard = types.InlineKeyboardMarkup()
        key_yes4 = types.InlineKeyboardButton(text='Да', callback_data='yes4')
        keyboard.add(key_yes4)
        key_no4 = types.InlineKeyboardButton(text='Нет', callback_data='no4')
        keyboard.add(key_no4)
        question = 'Нужен Wi-Fi/Bluetooth?'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "much2":
        bot.send_message(call.message.chat.id, answokfunc())
        if price2 !="<100$":
            price2="500-1000$"
        series1="i7"
        price3="200-300$"
        z=2
        keyboard = types.InlineKeyboardMarkup()
        key_yes4 = types.InlineKeyboardButton(text='Да', callback_data='yes4')
        keyboard.add(key_yes4)
        key_no4 = types.InlineKeyboardButton(text='Нет', callback_data='no4')
        keyboard.add(key_no4)
        question = 'Нужен Wi-Fi/Bluetooth?'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "toomuch2":
        bot.send_message(call.message.chat.id, answokfunc())
        if price2 !="<100$":
            price2=">1000$"
        series1="i9"
        price3="200-300$"
        z=3
        keyboard = types.InlineKeyboardMarkup()
        key_yes4 = types.InlineKeyboardButton(text='Да', callback_data='yes4')
        keyboard.add(key_yes4)
        key_no4 = types.InlineKeyboardButton(text='Нет', callback_data='no4')
        keyboard.add(key_no4)
        question = 'Нужен Wi-Fi/Bluetooth?'
        bot.send_message(call.from_user.id, text=question, reply_markup=keyboard)
    elif call.data == "yes3":
        wf3="Wi-Fi/Bluetooth – присутствуют"
        answ_message(call)
    elif call.data == "no3":
        wf3="Wi-Fi/Bluetooth – отсутствуют"
        answ_message(call)
    elif call.data == "yes4":
        wf3="Wi-Fi/Bluetooth – присутствуют"
        answ_message(call)
    elif call.data == "no4":
        wf3="Wi-Fi/Bluetooth – отсутствуют"
        answ_message(call)

bot.polling(none_stop=True, interval=0)
